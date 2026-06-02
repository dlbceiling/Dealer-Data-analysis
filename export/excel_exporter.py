from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import AnalysisResult


def export_analysis_result(result: AnalysisResult, output_path: str | Path) -> None:
    """导出分析结果到 Excel。"""
    output_path = Path(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        result.clean_detail.to_excel(writer, sheet_name="清洗后明细", index=False)
        result.product_summary.to_excel(writer, sheet_name="商品汇总", index=False)
        result.category_summary.to_excel(writer, sheet_name="营销分类汇总", index=False)
        result.honeycomb_summary.to_excel(writer, sheet_name="蜂窝板统计", index=False)
        result.category_margin_summary.to_excel(writer, sheet_name="营销分类毛利", index=False)
        result.missing_cost_summary.to_excel(writer, sheet_name="成本缺失清单", index=False)
        result.cost_exception_summary.to_excel(writer, sheet_name="成本异常清单", index=False)
        result.ignored_detail.to_excel(writer, sheet_name="已忽略数据", index=False)

        for sheet in writer.book.worksheets:
            _format_sheet(sheet)


def _format_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="E8F0FE")
    header_font = Font(bold=True, color="1F2937")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        width = min(max(max_length + 2, 12), 36)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
